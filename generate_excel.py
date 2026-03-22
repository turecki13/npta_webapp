"""
Skrypt generujący plik Excel z przykładowymi danymi dla panelu obozu tenisowego.
Uruchom: python generate_excel.py

Wygeneruje plik: dane_obozu.xlsx
z arkuszami:
  - participants    (uczestnicy)
  - days            (dni obozu)
  - schedule        (plan dnia – wspólny)
  - events          (wydarzenia specjalne)
    - personal_plans  (dodatkowe aktywności uczestników)
    - Grupa           (grupy treningowe)
    - Americano       (zapisy na americano)
    - Turniej tie breakowy
    - Turniej singlowy
    - Wycieczki
    - Grupa_wyjatki   (dzienne wyjątki dla grup)
"""

import pandas as pd

OUTPUT_FILE = "dane_obozu.xlsx"

# ── 1. UCZESTNICY ──────────────────────────────────────────────────────────────
# Kolumny:
#   name   – imię i nazwisko uczestnika (musi być unikalne)
#   active – TRUE / FALSE  (tylko TRUE pojawi się w aplikacji)
participants = pd.DataFrame([
    {"name": "Anna Kowalska",      "active": "TRUE"},
    {"name": "Bartek Nowak",       "active": "TRUE"},
    {"name": "Celina Wiśniewska",  "active": "TRUE"},
    {"name": "Dawid Zając",        "active": "FALSE"},  # nieaktywny – nie pojawi się w aplikacji
])

# ── 2. DNI OBOZU ───────────────────────────────────────────────────────────────
# Kolumny:
#   id       – unikalny identyfikator dnia (np. 1, 2, 3…)
#   title    – nagłówek dnia (np. "Niedziela 22.03")
#   subtitle – krótki opis dnia (opcjonalnie)
days = pd.DataFrame([
    {"id": "1", "title": "Niedziela 22.03",    "subtitle": "Dzień przyjazdu i pierwsze zajęcia"},
    {"id": "2", "title": "Poniedziałek 23.03", "subtitle": "Pełny dzień treningów"},
    {"id": "3", "title": "Wtorek 24.03",       "subtitle": "Turniej wewnętrzny"},
])

# ── 3. PLAN DNIA (wspólny harmonogram) ────────────────────────────────────────
# Kolumny:
#   day_id – id dnia (musi zgadzać się z kolumną id w arkuszu days)
#   order  – kolejność wyświetlania (liczba)
#   text   – treść wpisu w planie (np. "09:00 – Przyjazd i zakwaterowanie")
schedule = pd.DataFrame([
    {"day_id": "1", "order": 1, "text": "09:00 – Przyjazd i zakwaterowanie"},
    {"day_id": "1", "order": 2, "text": "11:00 – Rozgrzewka na korcie"},
    {"day_id": "1", "order": 3, "text": "19:00 – Kolacja powitalna"},
    {"day_id": "2", "order": 1, "text": "08:00 – Poranna rozgrzewka"},
    {"day_id": "2", "order": 2, "text": "09:00 – Trening z trenerem (2h)"},
    {"day_id": "2", "order": 3, "text": "17:00 – Sparingi"},
    {"day_id": "3", "order": 1, "text": "09:00 – Losowanie par turniejowych"},
    {"day_id": "3", "order": 2, "text": "10:00 – Turniej (faza grupowa)"},
    {"day_id": "3", "order": 3, "text": "16:00 – Finały i dekoracja"},
])

# ── 4. WYDARZENIA SPECJALNE ────────────────────────────────────────────────────
# Kolumny:
#   day_id – id dnia (musi zgadzać się z kolumną id w arkuszu days)
#   label  – treść komunikatu (wyświetlana jako niebieska ramka w aplikacji)
events = pd.DataFrame([
    {"day_id": "3", "label": "🏆 Turniej wewnętrzny – wszyscy uczestnicy"},
    {"day_id": "1", "label": "🍽️ Kolacja powitalna o 19:00"},
])

# ── 5. DODATKOWE AKTYWNOŚCI UCZESTNIKÓW ──────────────────────────────────────
# Kolumny:
#   participant_name – imię i nazwisko (musi zgadzać się z kolumną name w arkuszu participants)
#   day_id           – id dnia (musi zgadzać się z kolumną id w arkuszu days)
#   time             – godzina aktywności (opcjonalnie)
#   event_labels     – nazwa aktywności (np. "Wycieczka do Side")
#   note             – opcjonalna notatka
personal_plans = pd.DataFrame([
    {"participant_name": "Anna Kowalska",     "day_id": "1", "time": "14:00", "event_labels": "Wycieczka do Side", "note": "Zbiórka przy recepcji"},
    {"participant_name": "Anna Kowalska",     "day_id": "3", "time": "18:00", "event_labels": "Wieczór integracyjny", "note": "Strój sportowy"},
    {"participant_name": "Bartek Nowak",      "day_id": "3", "time": "18:00", "event_labels": "Wieczór integracyjny", "note": ""},
    {"participant_name": "Celina Wiśniewska", "day_id": "2", "time": "16:00", "event_labels": "Sesja fizjo", "note": "Gabinet 2"},
])

# ── 6. GRUPY TRENINGOWE ───────────────────────────────────────────────────────
# Kolumny:
#   participant_name – imię i nazwisko uczestnika
#   group            – nazwa grupy (np. "Grupa 1")
#   trainer          – trener prowadzący grupę
#   court            – domyślny kort grupy
#   training_time    – domyślna godzina treningu grupy
#   day_id           – opcjonalnie: id dnia; puste = grupa stała przez cały obóz
groups = pd.DataFrame([
    {"participant_name": "Anna Kowalska",     "group": "Grupa 1", "trainer": "Trener Tomek",    "court": "Kort 1", "training_time": "09:00", "day_id": ""},
    {"participant_name": "Bartek Nowak",      "group": "Grupa 1", "trainer": "Trener Tomek",    "court": "Kort 1", "training_time": "09:00", "day_id": ""},
    {"participant_name": "Celina Wiśniewska", "group": "Grupa 2", "trainer": "Trenerka Natalia", "court": "Kort 2", "training_time": "11:00", "day_id": ""},
])

# ── 7. AMERICANO ─────────────────────────────────────────────────────────────
# Kolumny:
#   participant_name – imię i nazwisko uczestnika
#   day_id           – opcjonalnie: id dnia; puste = obowiązuje dla wszystkich dni
americano = pd.DataFrame([
    {"participant_name": "Anna Kowalska", "day_id": "2"},
    {"participant_name": "Bartek Nowak",  "day_id": "2"},
])

# ── 8. TURNIEJ TIE BREAKOWY ──────────────────────────────────────────────────
tiebreak = pd.DataFrame([
    {"participant_name": "Celina Wiśniewska", "day_id": "3"},
])

# ── 9. TURNIEJ SINGLOWY ──────────────────────────────────────────────────────
singles = pd.DataFrame([
    {"participant_name": "Anna Kowalska", "day_id": "3"},
    {"participant_name": "Bartek Nowak",  "day_id": "3"},
])

# ── 10. WYCIECZKI ─────────────────────────────────────────────────────────────
trips = pd.DataFrame([
    {"participant_name": "Anna Kowalska", "day_id": "1"},
    {"participant_name": "Celina Wiśniewska", "day_id": "1"},
])

# ── 11. WYJĄTKI DZIENNE DLA GRUP ─────────────────────────────────────────────
# Kolumny:
#   day_id         – id dnia, dla którego obowiązuje wyjątek
#   group          – grupa, której dotyczy wyjątek (opcjonalnie, jeśli podasz participants)
#   participants   – opcjonalnie: skład na ten dzień (nazwiska po przecinku)
#   trainer        – opcjonalnie: trener na ten dzień
#   court          – opcjonalnie: kort na ten dzień
#   training_time  – opcjonalnie: godzina treningu na ten dzień
group_overrides = pd.DataFrame([
    {
        "day_id": "3",
        "group": "Grupa 1",
        "participants": "Anna Kowalska, Celina Wiśniewska",
        "trainer": "Trener Paweł",
        "court": "Kort 5",
        "training_time": "10:30",
    }
])

# ── ZAPIS DO EXCELA ────────────────────────────────────────────────────────────
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    participants.to_excel(writer,   sheet_name="participants",   index=False)
    days.to_excel(writer,           sheet_name="days",           index=False)
    schedule.to_excel(writer,       sheet_name="schedule",       index=False)
    events.to_excel(writer,         sheet_name="events",         index=False)
    personal_plans.to_excel(writer, sheet_name="personal_plans", index=False)
    groups.to_excel(writer,         sheet_name="Grupa",          index=False)
    americano.to_excel(writer,      sheet_name="Americano",      index=False)
    tiebreak.to_excel(writer,       sheet_name="Turniej tie breakowy", index=False)
    singles.to_excel(writer,        sheet_name="Turniej singlowy", index=False)
    trips.to_excel(writer,          sheet_name="Wycieczki", index=False)
    group_overrides.to_excel(writer, sheet_name="Grupa_wyjatki", index=False)

    # ── Formatowanie: szerokość kolumn ──────────────────────────────────────
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="1F7A4D")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]

        # Styl nagłówków
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Autodopasowanie szerokości kolumn
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)

print(f"✅ Plik '{OUTPUT_FILE}' został wygenerowany pomyślnie!")
print()
print("Arkusze:")
print("  • participants   – lista uczestników (name, active)")
print("  • days           – dni obozu (id, title, subtitle)")
print("  • schedule       – wspólny plan dnia (day_id, order, text)")
print("  • events         – wydarzenia specjalne (day_id, label)")
print("  • personal_plans – dodatkowe aktywności (participant_name, day_id, time, event_labels, note)")
print("  • Grupa          – grupy treningowe (participant_name, group, trainer, court, training_time, day_id opcjonalnie)")
print("  • Americano      – zapisy (participant_name, day_id opcjonalnie)")
print("  • Turniej tie breakowy – zapisy (participant_name, day_id opcjonalnie)")
print("  • Turniej singlowy     – zapisy (participant_name, day_id opcjonalnie)")
print("  • Wycieczki            – zapisy (participant_name, day_id opcjonalnie)")
print("  • Grupa_wyjatki        – wyjątki dnia (day_id, group, participants, trainer, court, training_time)")
